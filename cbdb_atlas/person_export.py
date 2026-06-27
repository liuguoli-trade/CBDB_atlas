"""Write Excel workbooks from export payloads (scheme B)."""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from cbdb_atlas.export_render import build_export_payload
from cbdb_atlas.store import CbdbStore


def _autosize_columns(ws) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        width = 10
        for cell in column_cells:
            if cell.value is None:
                continue
            width = max(width, min(len(str(cell.value)) + 2, 48))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def payload_to_workbook(payload: dict[str, Any]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    for sheet in payload["sheets"]:
        ws = wb.create_sheet(str(sheet["title"])[:31])
        ws.append(sheet["headers"])
        for row in sheet["rows"]:
            ws.append(row)
        _autosize_columns(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_person_workbook(store: CbdbStore, person_id: int) -> tuple[bytes, str]:
    payload = build_export_payload(store, person_id)
    content = payload_to_workbook(payload)
    return content, payload["filename"]
